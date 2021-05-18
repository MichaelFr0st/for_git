# openpyxl - открытие экселя, re - регулярки, fuzzywuzzy - Расстояние Левенштейна
from fuzzywuzzy import fuzz
from emoji import emojize
from fuzzywuzzy import process
import openpyxl
import re

#Создание пустых списков под заполнение данными из excel
inputData = []
storeData = []

#Список патернов для регулярных выражений
pattern1 = r'\b(\w+)[-\.,\\](\w+)\b'
pattern2 = r'(\w+)[-\.,\\_](\w+)[-\.,\\_](\w+)'
pattern3 = r'(\w+)[-\.,\\](\w+)[-\.,\\](\w+)[-\.,\\](\w+)'
pattern4 = r'(\w+)[-\.,\\](\w+)[-\.,\\](\w+)[-\.,\\](\w+)[-\.,\\](\w+)'

#Тут нужно будет вставить скомпилированые регулярки для многократного использования compile()
# reobj = re.compile("\b(\w+)[-\.,\\](\w+)\b")
obj = re.compile(r'[-\.,\\/_()#]')

#Открытие excel 1
bookone = openpyxl.open(r"/home/bender/Документы/rabs/fortest1.xlsx", read_only=True)
sheetone = bookone.active

#Открытие excel 2
booktwo = openpyxl.open(r"/home/bender/Документы/rabs/fortest2.xlsx", read_only=True)
sheettwo = booktwo.active

#Обработчик строк, разбивает строку по скомпиленой регулярке
def funk(stroka):
    if res := obj.split(stroka):
        return res
    else:
        print('to4no neverniy')
        return None

#Обработчик строк, возвращает словари
def kroba(stroka):
    if match := re.fullmatch(pattern1, stroka):
        raz, dva = match[1], match[2]
        #print('raz:', raz, 'dva:', dva)
        return {"raz" : raz, "dva" : dva}
    elif match := re.fullmatch(pattern2, stroka):
        raz, dva, tri = match[1], match[2], match[3]
        #print('raz:', raz, 'dva:', dva, 'tri:', tri)
        return {"raz" : raz, "dva" : dva, "tri" : tri}
    elif match := re.fullmatch(pattern3, stroka):
        raz, dva, tri, chetire, = match[1], match[2], match[3], match[4]
        #print('raz:', raz, 'dva:', dva, 'tri:', tri, 'chetire:', chetire)
        return {"raz" : raz, "dva" : dva, "tri" : tri, "chetire" : chetire}
    elif match := re.fullmatch(pattern4, stroka):
        raz, dva, tri, chetire, pat = match[1], match[2], match[3], match[4], match[5]
        #print('raz:', raz, 'dva:', dva, 'tri:', tri, 'chetire:', chetire, 'pat:', pat)
        return {"raz" : raz, "dva" : dva, "tri" : tri, "chetire": chetire, "pat" : pat}
    else:
        print(stroka, 'neverniy')
        return None

#Заполнение списка, на входе Имя списка, Имя таблицы, Режим
#Для заполнения id используются функции(режимы): kroba(более точное сравнение по патернам) или funk(тупа сплитую строку)
def zapolnator(spisok, table, rejim):
    if rejim == funk:
        for i, row in enumerate(range(1, table.max_row+1)):
            kod = table[row][0].value
            name = table[row][1].value
            spisok.append({"name": name, "id": funk(kod)})
    elif rejim == kroba:
        for i, row in enumerate(range(1, table.max_row+1)):
            kod = table[row][0].value
            name = table[row][1].value
            spisok.append({"name":name, "id":kroba(kod)})
    else:
        print('huinu delaesh')
        return None

#Вставка в значения в эксель(Названия листа, Данные для вставки, Нрмер колонки, Итератор списка(номер строки))
def ex_vstavka(list_name, strok, col_num, itr):
    value = str(strok)
    cell = list_name.cell(row = itr+1, column = col_num)
    cell.value = value

#Отрабатывает функция заполнения
zapolnator(inputData,sheetone,funk)
zapolnator(storeData,sheettwo, funk)

#Создание документа(таблицы) для финального результата
wb = openpyxl.Workbook()
sheet = wb['Sheet']

#Логика сравнения:
# 1-ое сравнение In[id]:Ex[id], запись IN[id] в 1 столбец, EX[id] запись в 4 столбец
# 2-ое сравнение In[id]:Ex[name], запись EX[name] в 5 столбец
# 3-е сравнение In[name]:Ex[name], запись IN[name] в 2 столбец
# 4-е сравнение In[kolvo]:Ex[kolvo], запись IN[kolvo] в 3 столбец, EX[kolvo] запись в 6 столбец
# 7 столбец нужен для цены за штуку(наша цена) для дальнейшего расчета формул и формирования предложения

#Сравнение списков и построчная запись в таблицу
for it in range(len(inputData)):
    kof = 0
    if troka:=inputData[it]['id']:
        ex_vstavka(sheet, inputData[it]['id'], 1, it)
        ex_vstavka(sheet, inputData[it]['name'],2, it)
        for ik in range(len(storeData)):
            b = fuzz.WRatio(inputData[it]['id'],storeData[ik]['id'])
            if b == 100:
                hroka = storeData[ik]['id']
                break
            elif b < 70:
                hroka = '\N{no entry}'
            elif b > kof:
                hroka = storeData[ik]['id']
            else:
                hroka = ''
                break
    ex_vstavka(sheet, hroka, 4, it)
    ex_vstavka(sheet, storeData[ik]['name'], 5, it)

#Сохранение документа
wb.save(r"/home/bender/Документы/res/example.xlsx")

print(emojize(":thumbs_up:"))
